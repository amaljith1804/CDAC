from tkinter import *
from openpyxl import *
# make sure the file on the folder
memo = load_workbook('memo.xlsx')
#activating the bill
bill = memo.active
#creating function for excel sheet
def excel():
    bill.cell(row = 1,column = 1).value = "No:94xxxxxxx6"
# inserting function to insert the parameters
def insert():
    if (date.get()=="" and  
        serial_no.get() =="" and
        customer_name.get() == "" and
        item1_name.get()== "" and
        item1_rate.get()== "" and
        item1_qty.get()=="" and
        item2_name.get() == "" and
		item2_qty.get() == "" and
		item2_rate.get() == "" and
		item3_name.get() == "" and
		item3_qty.get() == "" and
		item3_rate.get() == ""):
        print("empty input")
    else:
        current_row = bill.max_row
        current_column = bill.max_column
        bill.cell(row=5, column=6).value = date.get()
        bill.cell(row=5, column=2).value = serial_no.get()
        bill.cell(row=6,column=2).value = customer_name.get()
        bill.cell(row=10, column=2).value = item1_name.get()
        bill.cell(row=11, column=4).value = item1_qty.get()
        bill.cell(row=11, column=5).value = item1_rate.get()
        bill.cell(row=12, column=2).value = item2_name.get()
        bill.cell(row=13, column=4).value = item2_qty.get()
        bill.cell(row=13, column=5).value = item2_rate.get()
        bill.cell(row=14, column=2).value = item3_name.get()
        bill.cell(row=15, column=4).value = item3_qty.get()
        bill.cell(row=15, column=5).value = item3_rate.get()
        memo.save('memo.xlsx')
        date.focus_set()
        clear()
# parameter functions
# focus_set method to set the focus on the desired widget if and only if the master window is focused. 
def parameter1(event):
    date.focus_set()
def parameter2(event):
    serial_no.focus_set()
def parameter3(event):
    customer_name.focus_set()
def parameter4(event):
    item1_name.focus_set()
def parameter5(event):
    item1_qty.focus_set()
def parameter6(event):
    item1_rate.focus_set()
def parameter7(event):
	item2_name.focus_set() 
def parameter9(event):
	item2_qty.focus_set()
def parameter10(event):
	item2_rate.focus_set()
def parameter11(event): 
	item3_name.focus_set() 
def parameter12(event):
	item3_qty.focus_set()
def parameter13(event):
	item3_rate.focus_set()
#clear finction 
def clear():
    item1_name.delete(0,END)
    item1_rate.delete(0,END)
    item1_qty.delete(0,END)
    item2_name.delete(0, END)
    item2_rate.delete(0, END)
    item2_qty.delete(0,END)
    item3_name.delete(0,END)
    item3_qty.delete(0,END)
    item3_rate.delete(0,END)
    date.delete(0,END)
    serial_no.delete(0,END)
    customer_name.delete(0,END)

# driver code
if __name__ == "__main__":
    root = Tk()
    root.configure(background = 'green')
    root.title("Cash Memo")
    root.geometry("1000x500")
    excel()
    heading = Label(root, text = "AMAZON CASH BILL", bg = "light blue",foreground="blue")
    heading2 = Label(root, text="Cash Memo", bg="light green",foreground="blue")
    date = Label(root,text="Date",bg="light green",foreground="blue")
    serial_no = Label(root, text="Serial No", bg="light green",foreground="blue")
    customer_name = Label(root, text="Customer Name", bg="light green",foreground="blue")
    item1_name = Label(root, text="1. Item 1 Name ", bg="light green",foreground="blue")
    item1_rate = Label(root, text="Rate", bg="light green",foreground="blue")
    item1_qty = Label(root,text="Qty", bg="light green",foreground="blue")
    item2_name = Label(root, text="2. Item 2 Name ", bg="light green",foreground="blue")
    item2_qty = Label(root, text="Qty", bg="light green",foreground="blue")
    item2_rate = Label(root, text="Rate", bg="light green",foreground="blue")
    item3_name = Label(root, text="3. Item 3 Name", bg="light green",foreground="blue")
    item3_qty = Label(root, text="Qty", bg="light green",foreground="blue")
    item3_rate = Label(root, text="Rate", bg="light green",foreground="blue") 
    heading.grid(row=10,column=19)
    heading2.grid(row=11, column=19)
    date.grid(row=12, column=18)
    serial_no.grid(row=12, column=22)
    customer_name.grid(row=13, column=18)
    item1_name.grid(row=15, column=18)
    item1_qty.grid(row =15, column=20 )
    item1_rate.grid(row = 15,column=22)
    item2_name.grid(row=16, column=18)
    item2_qty.grid(row=16, column=20)
    item2_rate.grid(row=16, column=22)
    item3_name.grid(row=17, column=18)
    item3_qty.grid(row=17, column=20)
    item3_rate.grid(row=17, column=22)

    date= Entry(root)
    serial_no = Entry(root)
    customer_name = Entry(root)
    item1_name = Entry(root)
    item1_rate = Entry(root)
    item1_qty = Entry(root)
    item2_name = Entry(root)
    item2_qty = Entry(root)
    item2_rate = Entry(root)
    item3_name = Entry(root)
    item3_qty = Entry(root)
    item3_rate = Entry(root) 

    date.bind("<Return>",parameter1)
    serial_no.bind("<Return>",parameter2)
    customer_name.bind("<Return>",parameter3)
    item1_name.bind("<Return>",parameter4)
    item1_qty.bind("<Return>",parameter5)
    item1_rate.bind("<Return>",parameter6)
    item2_name.bind("<Return>", parameter7)
    item2_qty.bind("<Return>", parameter9)
    item2_rate.bind("<Return>", parameter10)
    item3_name.bind("<Return>", parameter11)
    item3_qty.bind("<Return>", parameter12)
    item3_rate.bind("<Return>", parameter13)

    date.grid(row = 12, column=19,ipadx="50")
    serial_no.grid(row=12,column=23,ipadx="1")
    customer_name.grid(row = 13,column=19,ipadx="50")
    item1_name.grid(row=15, column=19, ipadx="50")
    item1_rate.grid(row=15, column=23, ipadx="1")
    item1_qty.grid(row=15, column=21, ipadx="1")
    item2_name.grid(row=16, column=19, ipadx="50")
    item2_qty.grid(row=16, column=21, ipadx="1")
    item2_rate.grid(row=16, column=23, ipadx="1")
    item3_name.grid(row=17, column=19, ipadx="50")
    item3_qty.grid(row=17, column=21, ipadx="1")
    item3_rate.grid(row=17, column=23, ipadx="1")

    excel()
    submit = Button(root,text = "Submit", fg ="Black",bg = "Red", command = insert)
    submit.grid(row=30,column=19)
    root.mainloop() 

